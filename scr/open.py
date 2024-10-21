
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def df_to_excel_adaptive(df, output_path, sheet_name='Sheet1'):
    """
    Convert a pandas DataFrame to an Excel file with adaptive column widths.

    :param df: pandas DataFrame to convert
    :param output_path: path to save the Excel file
    :param sheet_name: name of the sheet in Excel (default is 'Sheet1')
    """
    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Drop rows with NaN values in 'PRICE.1' column
    df = df.dropna(subset=['PRICE.1'])

    # Keep only specified columns
    columns_to_keep = [
        'SHOP_SKU', 'OFFER', 'MAIN_PRICE', 'MERCH_PRICE_WITH_PROMOS',
        'PRICE_GREEN_THRESHOLD', 'PRICE_RED_THRESHOLD', 'PRICE_WITH_PROMOS',
        'SHOP_WITH_BEST_PRICE_ON_MARKET', 'PRICE.1'
    ]
    df = df[columns_to_keep]

    # Convert DataFrame to rows
    rows = dataframe_to_rows(df, index=False, header=True)

    # Write data to worksheet
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Make the header row bold
            if r_idx == 1:
                cell.font = Font(bold=True)

    # Calculate and set column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(output_path)
    print(f"Excel file saved to {output_path}")

# Read the CSV file
df = pd.read_csv('report_old.csv')

# Call the function to create Excel file
df_to_excel_adaptive(df, 'output.xlsx')
